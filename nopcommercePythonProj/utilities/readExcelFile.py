import openpyxl
from utilities.ReadConfigFile import ReadConfigFile

  
class readExcelFile:
      
    filePath = ReadConfigFile.getDataFilePath()
     
    def __init__(self):
        self.workBook = openpyxl().load_workbook(self.filePath)
         
    def readExcelFileSheet(self,sheetName):
        self.sheet = self.workBook[sheetName]
        return self.sheet
         
    def getRowCount1(self,sheetName):
        self.rowMaxCount= self.workBook[sheetName]
        self.rowMaxCount1.max_row()
        return self.rowMaxCount1
#     
#     def getColumnCount(self,sheetName):
#         self.columnMaxCount = self.workBook[sheetName].max_column()
#         return self.columnMaxCount
     
    def readCellData(self,sheetName,rowNum,columnNum):
        self.cell = self.workBook[sheetName].cell(rowNum,columnNum).value
        return self.cell
#     
#     def writeData(self,sheetName,rowNum,columnNum,data):
#         self.write = self.workBook[sheetName].cell(rowNum,columnNum).Value=data
#         self.workBook.save()

def getRowCout(file,sheetName):
#     workbook = openpyxl.load_workbook(file)
#     sheet = workbook[sheetName]
#     return (sheet.max_row)
    return openpyxl.load_workbook(file)[sheetName].max_row

def getColCount(file,sheetName):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    return (sheet.Max_column)

def readData(file,sheetName,rownum,colnum):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    return sheet.cell(row=rownum,column=colnum).value

#     